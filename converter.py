#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Converter module for transforming test cases to CSV and Excel formats.
"""

import os
import csv
from typing import Dict, List, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from loguru import logger


class TestCaseConverter:
    """Converter for transforming test cases to CSV and Excel formats."""

    # Common test case fields with order preservation
    TEST_CASE_FIELDS = [
        "ID", "Name", "Desc", "Pre-conditions", "Test Steps", "Expected Result", 
        "Actual Result", "Test Data", "Priority", "Severity", "Status",
        "Environment", "Tested By", "Date", "Comments/Notes"
    ]

    def __init__(self, output_dir: str = "output"):
        """
        Initialize the converter.

        Args:
            output_dir: Directory where output files will be saved.
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def convert_to_csv(self, test_cases: Dict[str, List[Dict[str, Any]]], force: bool = False) -> Dict[str, str]:
        """
        Convert test cases to CSV files.

        Args:
            test_cases: Dictionary with test case file names as keys and lists of test case dictionaries as values.
            force: Whether to overwrite existing files without asking.

        Returns:
            Dictionary mapping file names to output paths.
        """
        output_files = {}
        
        for file_name, cases in test_cases.items():
            if not cases:
                logger.warning(f"No test cases to convert for {file_name}")
                continue
                
            # Create output file path
            base_name = Path(file_name).stem
            output_path = os.path.join(self.output_dir, f"{base_name}.csv")
            
            # Check if file exists
            if os.path.exists(output_path) and not force:
                response = input(f"File {output_path} already exists. Overwrite? (y/n): ")
                if response.lower() != 'y':
                    logger.info(f"Skipping {output_path}")
                    continue
            
            try:
                with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                    fieldnames = self.TEST_CASE_FIELDS
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    
                    writer.writeheader()
                    for case in cases:
                        # Normalize keys (handle case sensitivity)
                        normalized_case = {}
                        for field in fieldnames:
                            # Try exact match first
                            if field in case:
                                normalized_case[field] = case[field]
                            else:
                                # Try case-insensitive match
                                field_lower = field.lower()
                                found = False
                                for key in case:
                                    if key.lower() == field_lower:
                                        normalized_case[field] = case[key]
                                        found = True
                                        break
                                if not found:
                                    normalized_case[field] = ""
                        
                        writer.writerow(normalized_case)
                
                logger.info(f"Created CSV file: {output_path}")
                output_files[file_name] = output_path
                
            except Exception as e:
                logger.error(f"Error creating CSV file {output_path}: {str(e)}")
        
        return output_files

    def convert_to_excel(self, test_cases: Dict[str, List[Dict[str, Any]]], force: bool = False) -> Optional[str]:
        """
        Convert all test cases to a single Excel file with multiple sheets.

        Args:
            test_cases: Dictionary with test case file names as keys and lists of test case dictionaries as values.
            force: Whether to overwrite existing files without asking.

        Returns:
            Path to the created Excel file, or None if creation failed.
        """
        if not test_cases:
            logger.warning("No test cases to convert to Excel")
            return None
            
        excel_path = os.path.join(self.output_dir, "test_cases.xlsx")
        
        # Check if file exists
        if os.path.exists(excel_path) and not force:
            response = input(f"File {excel_path} already exists. Overwrite? (y/n): ")
            if response.lower() != 'y':
                logger.info(f"Skipping Excel file creation")
                return None
        
        try:
            workbook = openpyxl.Workbook()
            # Remove the default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
            for file_name, cases in test_cases.items():
                if not cases:
                    continue
                    
                # Create a sheet for each file
                sheet_name = Path(file_name).stem[:31]  # Excel sheet names are limited to 31 chars
                sheet = workbook.create_sheet(sheet_name)
                
                # Add header row
                for col_idx, field in enumerate(self.TEST_CASE_FIELDS, start=1):
                    cell = sheet.cell(row=1, column=col_idx, value=field)
                    # Style header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Add test case data
                for row_idx, case in enumerate(cases, start=2):
                    for col_idx, field in enumerate(self.TEST_CASE_FIELDS, start=1):
                        # Try exact match first
                        value = ""
                        if field in case:
                            value = case[field]
                        else:
                            # Try case-insensitive match
                            field_lower = field.lower()
                            for key in case:
                                if key.lower() == field_lower:
                                    value = case[key]
                                    break
                        
                        cell = sheet.cell(row=row_idx, column=col_idx, value=str(value))
                        cell.alignment = Alignment(wrap_text=True)
                
                # Auto-adjust column widths
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    
                    # Limit column width to a reasonable size
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column].width = adjusted_width
            
            workbook.save(excel_path)
            logger.info(f"Created Excel file: {excel_path}")
            return excel_path
            
        except Exception as e:
            logger.error(f"Error creating Excel file {excel_path}: {str(e)}")
            return None
