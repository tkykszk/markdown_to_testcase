#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Tests for the TestCaseConverter class.
"""

import os
import csv
import pytest
import tempfile
import sys
from pathlib import Path
import openpyxl

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from converter import TestCaseConverter


@pytest.fixture
def converter():
    """Create a TestCaseConverter instance with a temporary output directory."""
    with tempfile.TemporaryDirectory() as temp_dir:
        yield TestCaseConverter(output_dir=temp_dir)


@pytest.fixture
def sample_test_cases():
    """Sample test cases for testing."""
    return {
        "test_file1.md": [
            {
                "ID": "TC001",
                "Name": "Test Case 1",
                "Desc": "Description for test case 1",
                "Test Steps": "Steps for test case 1",
                "Expected Result": "Expected result 1",
                "Priority": "High"
            },
            {
                "ID": "TC002",
                "Name": "Test Case 2",
                "Desc": "Description for test case 2",
                "Test Steps": "Steps for test case 2",
                "Expected Result": "Expected result 2",
                "Priority": "Medium"
            }
        ],
        "test_file2.md": [
            {
                "ID": "TC101",
                "Name": "Test Case 101",
                "Desc": "Description for test case 101",
                "Test Steps": "Steps for test case 101",
                "Expected Result": "Expected result 101",
                "Priority": "Low"
            }
        ]
    }


def test_convert_to_csv(converter, sample_test_cases, monkeypatch):
    """Test converting test cases to CSV files."""
    # Mock the input function to always return 'y'
    monkeypatch.setattr('builtins.input', lambda _: 'y')
    
    # Convert to CSV
    result = converter.convert_to_csv(sample_test_cases, force=False)
    
    # Check that the output files were created
    assert len(result) == 2
    assert "test_file1.md" in result
    assert "test_file2.md" in result
    
    # Check the content of the first CSV file
    csv_path = result["test_file1.md"]
    assert os.path.exists(csv_path)
    
    with open(csv_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        rows = list(reader)
        
        assert len(rows) == 2
        assert rows[0]["ID"] == "TC001"
        assert rows[0]["Name"] == "Test Case 1"
        assert rows[1]["ID"] == "TC002"
        assert rows[1]["Name"] == "Test Case 2"
    
    # Check the content of the second CSV file
    csv_path = result["test_file2.md"]
    assert os.path.exists(csv_path)
    
    with open(csv_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        rows = list(reader)
        
        assert len(rows) == 1
        assert rows[0]["ID"] == "TC101"
        assert rows[0]["Name"] == "Test Case 101"


def test_convert_to_excel(converter, sample_test_cases, monkeypatch):
    """Test converting test cases to an Excel file."""
    # Mock the input function to always return 'y'
    monkeypatch.setattr('builtins.input', lambda _: 'y')
    
    # Convert to Excel
    result = converter.convert_to_excel(sample_test_cases, force=False)
    
    # Check that the output file was created
    assert result is not None
    assert os.path.exists(result)
    
    # Check the content of the Excel file
    workbook = openpyxl.load_workbook(result)
    
    # Check that there are two sheets
    assert len(workbook.sheetnames) == 2
    assert "test_file1" in workbook.sheetnames
    assert "test_file2" in workbook.sheetnames
    
    # Check the content of the first sheet
    sheet = workbook["test_file1"]
    assert sheet.cell(1, 1).value == "ID"  # Header
    assert sheet.cell(1, 2).value == "Name"  # Header
    assert sheet.cell(2, 1).value == "TC001"
    assert sheet.cell(2, 2).value == "Test Case 1"
    assert sheet.cell(3, 1).value == "TC002"
    assert sheet.cell(3, 2).value == "Test Case 2"
    
    # Check the content of the second sheet
    sheet = workbook["test_file2"]
    assert sheet.cell(1, 1).value == "ID"  # Header
    assert sheet.cell(1, 2).value == "Name"  # Header
    assert sheet.cell(2, 1).value == "TC101"
    assert sheet.cell(2, 2).value == "Test Case 101"


def test_convert_to_csv_force(converter, sample_test_cases):
    """Test converting test cases to CSV files with force=True."""
    # Convert to CSV with force=True
    result = converter.convert_to_csv(sample_test_cases, force=True)
    
    # Check that the output files were created
    assert len(result) == 2
    
    # Test overwriting the same files
    result2 = converter.convert_to_csv(sample_test_cases, force=True)
    
    # Check that the output files were created again
    assert len(result2) == 2


def test_convert_empty_test_cases(converter):
    """Test converting empty test cases."""
    # Convert empty test cases to CSV
    result_csv = converter.convert_to_csv({})
    assert result_csv == {}
    
    # Convert empty test cases to Excel
    result_excel = converter.convert_to_excel({})
    assert result_excel is None
